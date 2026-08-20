"""Acrobatでxlsx化したサロンレポートの読み取り。

PDF直読み(report_pdf.py)が本命だが、Acrobat変換に慣れている間の入力経路として、
またPDFの版が変わって直読みが崩れたときの逃げ道として残してある。
セル格子に変換したあとの解析は report_grid.py と共通。
"""

from __future__ import annotations

from openpyxl import load_workbook

from .report_grid import COUPON_SECTION_MARKER, SalonReport, parse_grid


def _sheet_to_grid(worksheet) -> list[list[object]]:
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def parse_report_xlsx(path: str) -> SalonReport:
    """Acrobat変換済みのサロンレポートxlsxを読み取る。"""
    workbook = load_workbook(path, data_only=True, read_only=True)
    report = SalonReport(source=path)
    try:
        parsed_any = False
        for worksheet in workbook.worksheets:
            if parse_grid(_sheet_to_grid(worksheet), report):
                parsed_any = True
                break
        if not parsed_any and not report.warnings:
            report.warnings.append(
                f'どのシートにも「{COUPON_SECTION_MARKER}」が見つかりませんでした。'
                'Acrobatの変換範囲にランキングのページが含まれているか確認してください。'
            )
    finally:
        workbook.close()
    return report
