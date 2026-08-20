"""照合結果をxlsxに書き出す。

シート構成は「そのまま眺めて判断できること」を優先している。
    照合結果        掲載クーポン全件 × 予約数。予約が多い順。
    予約ゼロ        掲載中なのに予約実績が無いもの。差し替え候補。
    未照合の予約    予約はあるが現在の掲載に無いもの。停止/改名したクーポン。
    要確認          完全一致しなかった対応。人の目で確かめる行だけを集めたもの。
    実行情報        いつ・どのファイル・どこから取ったかの記録。
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .pipeline import ReconcileResult

_HEADER_FILL = PatternFill('solid', fgColor='1F3864')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_ZERO_FILL = PatternFill('solid', fgColor='FCE4E4')
_FUZZY_FILL = PatternFill('solid', fgColor='FFF2CC')
_MAX_WIDTH = 60
_METHOD_LABEL = {'exact': '完全一致', 'fuzzy': 'あいまい一致', 'unmatched': '対応なし'}


def _write_header(worksheet, headers: list[str]) -> None:
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'


def _fit_columns(worksheet, headers: list[str]) -> None:
    for index, header in enumerate(headers, 1):
        longest = len(header)
        for cell in worksheet[get_column_letter(index)][1:]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            _MAX_WIDTH, max(8, longest + 2)
        )


def _sheet(workbook: Workbook, title: str, headers: list[str], first: bool = False):
    worksheet = workbook.active if first else workbook.create_sheet()
    worksheet.title = title
    _write_header(worksheet, headers)
    return worksheet


def _build_match_sheet(workbook: Workbook, result: ReconcileResult) -> None:
    labels = result.issue_labels
    headers = (
        ['掲載順', '客区分', 'クーポン名(掲載)', '価格']
        + labels
        + ['予約数合計', '照合方法', '一致率', 'クーポン名(レポート側)']
    )
    worksheet = _sheet(workbook, '照合結果', headers, first=True)

    for row in result.rows:
        worksheet.append(
            [row.order, row.customer_type, row.listed_name, row.price]
            + [row.monthly.get(label, 0) for label in labels]
            + [
                row.total,
                _METHOD_LABEL.get(row.method, row.method),
                round(row.score, 4) if row.score is not None else None,
                row.reserved_name,
            ]
        )

    for excel_row, row in enumerate(result.rows, 2):
        if row.total == 0:
            fill = _ZERO_FILL
        elif row.method == 'fuzzy':
            fill = _FUZZY_FILL
        else:
            continue
        for col in range(1, len(headers) + 1):
            worksheet.cell(row=excel_row, column=col).fill = fill
    _fit_columns(worksheet, headers)


def _build_zero_sheet(workbook: Workbook, result: ReconcileResult) -> None:
    labels = result.issue_labels
    headers = ['掲載順', '客区分', 'クーポン名(掲載)', '価格'] + labels + ['予約数合計']
    worksheet = _sheet(workbook, '予約ゼロ', headers)
    for row in sorted(result.zero_reservation_rows, key=lambda r: r.order):
        worksheet.append(
            [row.order, row.customer_type, row.listed_name, row.price]
            + [row.monthly.get(label, 0) for label in labels]
            + [row.total]
        )
    _fit_columns(worksheet, headers)


def _build_orphan_sheet(workbook: Workbook, result: ReconcileResult) -> None:
    labels = result.issue_labels
    headers = ['クーポン名(レポート側)'] + labels + ['予約数合計']
    worksheet = _sheet(workbook, '未照合の予約', headers)
    for orphan in result.orphans:
        worksheet.append(
            [orphan.name] + [orphan.monthly.get(label, 0) for label in labels] + [orphan.total]
        )
    _fit_columns(worksheet, headers)


def _build_review_sheet(workbook: Workbook, result: ReconcileResult) -> None:
    headers = ['一致率', '掲載順', 'クーポン名(掲載)', 'クーポン名(レポート側)', '予約数合計']
    worksheet = _sheet(workbook, '要確認', headers)
    for row in sorted(result.needs_review_rows, key=lambda r: r.score or 0):
        worksheet.append(
            [
                round(row.score, 4) if row.score is not None else None,
                row.order,
                row.listed_name,
                row.reserved_name,
                row.total,
            ]
        )
    _fit_columns(worksheet, headers)


def _build_info_sheet(workbook: Workbook, result: ReconcileResult) -> None:
    headers = ['項目', '内容']
    worksheet = _sheet(workbook, '実行情報', headers)
    report = result.report
    rows: list[tuple[str, object]] = [
        ('実行日時', result.executed_at),
        ('レポートファイル', report.source if report else ''),
        ('貴店名', report.salon_name if report else ''),
        ('ご掲載月号', report.issue if report else ''),
        ('データ最終更新日', report.updated_at if report else ''),
        ('集計対象の月号', ' / '.join(result.issue_labels)),
        ('掲載クーポンの取得元', result.coupon_source),
        ('掲載クーポン件数', len(result.rows)),
        ('レポートの予約データ件数', len(report.reservations) if report else 0),
        ('完全一致', result.exact_count),
        ('あいまい一致(要確認)', result.fuzzy_count),
        ('予約ゼロの掲載クーポン', len(result.zero_reservation_rows)),
        ('未照合の予約データ', len(result.orphans)),
        ('予約数の総計', sum(r.total for r in result.rows)),
    ]
    for item, value in rows:
        worksheet.append([item, value])
    if result.warnings:
        worksheet.append([])
        worksheet.append(['警告', ''])
        for warning in result.warnings:
            worksheet.append(['', warning])
    _fit_columns(worksheet, headers)


def write_workbook(result: ReconcileResult, output_path: str) -> str:
    """照合結果をxlsxとして保存し、保存先のパスを返す。"""
    workbook = Workbook()
    _build_match_sheet(workbook, result)
    _build_zero_sheet(workbook, result)
    _build_orphan_sheet(workbook, result)
    _build_review_sheet(workbook, result)
    _build_info_sheet(workbook, result)
    workbook.save(output_path)
    return output_path
