"""レポート読み取り〜照合〜出力までの通しの検証。"""

from openpyxl import load_workbook

from fixtures import (
    ISSUE_LABELS,
    LISTED_MONTHS,
    NEWLY_LISTED,
    RESERVED_COUPONS,
    SAMPLE_COUPONS,
    build_coupon_html,
    write_report_xlsx,
)
from hpb_coupon.excel_out import write_workbook
from hpb_coupon.pipeline import reconcile
from hpb_coupon.scrape import parse_coupon_html


def _write_inputs(tmp_path):
    report = write_report_xlsx(str(tmp_path / 'report.xlsx'))
    html = tmp_path / 'coupon.html'
    html.write_text(build_coupon_html(), encoding='utf-8')
    return report, str(html)


def test_クーポンHTMLからは価格と掲載順が取れる():
    coupons = parse_coupon_html(build_coupon_html())
    assert len(coupons) == len(SAMPLE_COUPONS)
    assert coupons[0].order == 1
    assert coupons[0].customer_type == SAMPLE_COUPONS[0][0]
    assert coupons[0].name == SAMPLE_COUPONS[0][1]
    assert coupons[0].price == 8100
    assert coupons[0].conditions == '他券併用不可1'


def test_掲載クーポン基準で予約数が並ぶ(tmp_path):
    report, html = _write_inputs(tmp_path)
    result = reconcile(report_path=report, coupon_html_paths=[html])

    assert len(result.rows) == len(SAMPLE_COUPONS)
    assert result.issue_labels == ISSUE_LABELS
    # 表記ゆれはすべて完全一致として吸収される
    assert result.exact_count == len(RESERVED_COUPONS)
    assert result.fuzzy_count == 0
    # 予約が多い順に並ぶ
    assert [r.total for r in result.rows] == sorted(
        [r.total for r in result.rows], reverse=True
    )
    assert result.rows[0].total == sum(RESERVED_COUPONS[0][2])
    assert result.rows[0].price == 8100


def test_予約ゼロの掲載クーポンが抽出される(tmp_path):
    report, html = _write_inputs(tmp_path)
    result = reconcile(report_path=report, coupon_html_paths=[html])
    zero = result.zero_reservation_rows
    assert len(zero) == 1
    assert zero[0].listed_name.startswith('ヘッドスパ15分')
    assert zero[0].method == 'unmatched'


def test_レポートだけでも照合できる(tmp_path):
    """サロンURLもHTMLも渡さない=通信なしの経路。"""
    report = write_report_xlsx(str(tmp_path / 'report.xlsx'))
    result = reconcile(report_path=report)
    assert 'レポート内' in result.coupon_source
    assert len(result.rows) == len(SAMPLE_COUPONS)
    assert result.exact_count == len(RESERVED_COUPONS)
    # レポート内の一覧には価格が無い
    assert all(row.price is None for row in result.rows)


def test_掲載していないクーポンの予約は未照合に回る(tmp_path):
    report, html = _write_inputs(tmp_path)
    # 1件を掲載一覧から外す = 掲載を止めたクーポンの予約が残っている状態
    trimmed = build_coupon_html().replace(SAMPLE_COUPONS[1][1], 'まったく別の名前のクーポン')
    other = tmp_path / 'trimmed.html'
    other.write_text(trimmed, encoding='utf-8')

    result = reconcile(report_path=report, coupon_html_paths=[str(other)])
    assert len(result.orphans) == 1
    assert result.orphans[0].name.startswith('カット+縮毛矯正')
    assert result.orphans[0].total == sum(RESERVED_COUPONS[1][2])


def test_出力xlsxのシート構成と件数(tmp_path):
    report, html = _write_inputs(tmp_path)
    result = reconcile(report_path=report, coupon_html_paths=[html])
    output = write_workbook(result, str(tmp_path / 'out.xlsx'))

    workbook = load_workbook(output)
    assert workbook.sheetnames == ['照合結果', '予約ゼロ', '未照合の予約', '要確認', '実行情報']

    match_sheet = workbook['照合結果']
    headers = [cell.value for cell in match_sheet[1]]
    assert headers[:4] == ['掲載順', '客区分', 'クーポン名(掲載)', '価格']
    for label in ISSUE_LABELS:
        assert label in headers
    assert match_sheet.max_row == len(SAMPLE_COUPONS) + 1
    assert workbook['予約ゼロ'].max_row == 2

    info = {row[0]: row[1] for row in workbook['実行情報'].iter_rows(min_row=2, values_only=True)}
    assert info['完全一致'] == len(RESERVED_COUPONS)
    assert info['予約数の総計'] == sum(sum(c[2]) for c in RESERVED_COUPONS)


def test_掲載履歴から改名を検出する(tmp_path):
    """06月号の誤植が07月号で直された、という想定のクーポンを追えるか。"""
    report, html = _write_inputs(tmp_path)
    result = reconcile(report_path=report, coupon_html_paths=[html])

    renamed = [r for r in result.rows if r.listing_status == '改名']
    assert len(renamed) == 1
    row = renamed[0]
    assert row.listed_name.startswith('カット+縮毛矯正')
    assert '210000' in row.previous_name  # 06月号の誤植のほう
    assert row.previous_month == LISTED_MONTHS[0]
    assert row.rename_score is not None and row.rename_score >= 0.75
    # 改名前まで遡るので、初掲載は06月号になる
    assert row.first_listed_month == LISTED_MONTHS[0]


def test_途中から載り始めたクーポンは新規掲載になる(tmp_path):
    report, html = _write_inputs(tmp_path)
    result = reconcile(report_path=report, coupon_html_paths=[html])

    newly = [r for r in result.rows if r.listing_status == '新規掲載']
    assert [r.listed_name for r in newly] == [NEWLY_LISTED]
    assert newly[0].first_listed_month == LISTED_MONTHS[1]


def test_ずっと載っているクーポンは継続になる(tmp_path):
    report, html = _write_inputs(tmp_path)
    result = reconcile(report_path=report, coupon_html_paths=[html])

    kept = [r for r in result.rows if r.listing_status == '継続']
    assert len(kept) == len(SAMPLE_COUPONS) - 2  # 改名1件と新規掲載1件を除く
    assert all(r.first_listed_month == LISTED_MONTHS[0] for r in kept)


def test_改名の対応づけは1対1になる(tmp_path):
    """1つの旧名が複数のクーポンの旧名として使い回されないこと。"""
    report, html = _write_inputs(tmp_path)
    result = reconcile(report_path=report, coupon_html_paths=[html])

    previous = [r.previous_name for r in result.rows if r.previous_name]
    assert len(previous) == len(set(previous))


def test_未照合の予約に最終掲載月号と改名候補が付く(tmp_path):
    report, html = _write_inputs(tmp_path)
    # 改名されたクーポンを掲載一覧から外すと、その予約が未照合として残る
    trimmed = build_coupon_html().replace(SAMPLE_COUPONS[1][1], '掲載を止めた別のクーポン')
    other = tmp_path / 'trimmed.html'
    other.write_text(trimmed, encoding='utf-8')

    result = reconcile(report_path=report, coupon_html_paths=[str(other)])
    assert len(result.orphans) == 1
    orphan = result.orphans[0]
    assert orphan.last_listed_month == LISTED_MONTHS[1]
    assert orphan.rename_candidate == ''  # 似た掲載クーポンが無いので候補なし


def test_出力に掲載履歴の列が入る(tmp_path):
    report, html = _write_inputs(tmp_path)
    result = reconcile(report_path=report, coupon_html_paths=[html])
    output = write_workbook(result, str(tmp_path / 'out.xlsx'))

    workbook = load_workbook(output)
    headers = [cell.value for cell in workbook['照合結果'][1]]
    for column in ['掲載状況', '初掲載月号', '旧クーポン名', '旧名の月号', '改名の一致率']:
        assert column in headers

    orphan_headers = [cell.value for cell in workbook['未照合の予約'][1]]
    for column in ['最終掲載月号', '改名候補(現在の掲載)', '改名の一致率']:
        assert column in orphan_headers

    info = {row[0]: row[1] for row in workbook['実行情報'].iter_rows(min_row=2, values_only=True)}
    assert info['改名を検出したクーポン'] == 1
    assert info['掲載履歴が取れた月号'] == ' / '.join(LISTED_MONTHS)
